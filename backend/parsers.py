from __future__ import annotations

import csv
import io
import json
import re
from html import unescape
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".csv",
    ".json",
    ".html",
    ".htm",
    ".xml",
    ".log",
}


def extract_text(file_name: str, content_type: str | None, content: bytes) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".txt", ".md", ".markdown", ".log"}:
        return content.decode("utf-8", errors="ignore")
    if suffix == ".csv":
        return _extract_csv(content)
    if suffix == ".json":
        return _extract_json(content)
    if suffix in {".html", ".htm", ".xml"}:
        return _strip_markup(content.decode("utf-8", errors="ignore"))
    if suffix == ".pdf" or content_type == "application/pdf":
        return _extract_pdf(content)
    if suffix == ".docx":
        return _extract_docx(content)
    if suffix == ".xlsx":
        return _extract_xlsx(content)
    return content.decode("utf-8", errors="ignore")


def split_text(text: str, chunk_size: int, overlap: int) -> list[str]:
    normalized = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not normalized:
        return []

    paragraphs = [part.strip() for part in normalized.split("\n\n") if part.strip()]
    chunks: list[str] = []
    current = ""

    for paragraph in paragraphs:
        candidate = f"{current}\n\n{paragraph}".strip() if current else paragraph
        if len(candidate) <= chunk_size:
            current = candidate
            continue
        if current:
            chunks.append(current)
        if len(paragraph) <= chunk_size:
            current = paragraph
            continue
        start = 0
        while start < len(paragraph):
            end = min(len(paragraph), start + chunk_size)
            chunk = paragraph[start:end].strip()
            if chunk:
                chunks.append(chunk)
            if end >= len(paragraph):
                break
            start = max(end - overlap, start + 1)
        current = ""

    if current:
        chunks.append(current)

    return chunks


def _extract_csv(content: bytes) -> str:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    return "\n".join(" | ".join(cell.strip() for cell in row) for row in reader)


def _extract_json(content: bytes) -> str:
    raw = content.decode("utf-8", errors="ignore")
    try:
        parsed = json.loads(raw)
        flattened = _extract_farm_records_json(parsed)
        if flattened:
            return flattened
        return json.dumps(parsed, ensure_ascii=False, indent=2)
    except json.JSONDecodeError:
        return raw


def _extract_farm_records_json(parsed: object) -> str | None:
    if isinstance(parsed, list):
        records = parsed
    elif isinstance(parsed, dict) and isinstance(parsed.get("records"), list):
        records = parsed["records"]
    else:
        return None

    parts: list[str] = []

    for record in records:
        if not isinstance(record, dict):
            continue

        kind = record.get("kind")
        name = record.get("display_name") or record.get("name")
        if kind not in {"farm", "producer", "market", "shop"} or not isinstance(name, str) or not name.strip():
            continue

        products = [str(value).strip() for value in record.get("products", []) if value]
        categories = [str(value).strip() for value in record.get("categories", []) if value]
        certifications = [str(value).strip() for value in record.get("certifications", []) if value]

        source_refs: list[str] = []
        for source in record.get("sources", []):
            if isinstance(source, dict) and source.get("ref"):
                source_refs.append(str(source["ref"]).strip())

        lines = [
            f"ID: {record.get('id', '')}".strip(),
            f"Name: {name}",
            f"Type: {kind}",
            f"County: {record.get('county')}" if record.get("county") else "",
            f"Municipality: {record.get('municipality')}" if record.get("municipality") else "",
            f"Latitude: {record.get('lat')}" if record.get("lat") is not None else "",
            f"Longitude: {record.get('lng')}" if record.get("lng") is not None else "",
            f"Products: {', '.join(products)}" if products else "",
            f"Categories: {', '.join(categories)}" if categories else "",
            f"Certifications: {', '.join(certifications)}" if certifications else "",
            f"Description: {record.get('description')}" if record.get("description") else "",
            f"Contact email: {record['contact'].get('email')}" if isinstance(record.get("contact"), dict) and record["contact"].get("email") else "",
            f"Contact phone: {record['contact'].get('phone')}" if isinstance(record.get("contact"), dict) and record["contact"].get("phone") else "",
            f"Website: {record['contact'].get('website')}" if isinstance(record.get("contact"), dict) and record["contact"].get("website") else "",
            f"Sources: {' | '.join(source_refs)}" if source_refs else "",
        ]
        text = "\n".join(line for line in lines if line)
        if text:
            parts.append(text)

    if not parts:
        return None

    return "\n\n---\n\n".join(parts)


def _extract_pdf(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    parts: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            parts.append(page_text.strip())
    return "\n\n".join(parts)


def _extract_docx(content: bytes) -> str:
    doc = Document(io.BytesIO(content))
    paragraphs = [paragraph.text.strip() for paragraph in doc.paragraphs if paragraph.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(content: bytes) -> str:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                rows.append(" | ".join(values))
        if rows:
            parts.append(f"# Sheet: {sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _strip_markup(text: str) -> str:
    without_scripts = re.sub(r"<(script|style)[^>]*>.*?</\\1>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    without_tags = re.sub(r"<[^>]+>", " ", without_scripts)
    return re.sub(r"\s+", " ", unescape(without_tags)).strip()
